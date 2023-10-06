################################################################################################################
# Reading data from excel sheet

from openpyxl import load_workbook
import pandas as pd
import numpy as np
import array
import utils

from pyautocad import Autocad, APoint
from math import *
import win32com.client
import pythoncom
# from math import pi

# Initial Data
Start_Column = 6
Start_Row = 11
End_Column = 21
End_Row = 37

Shaft_Scale = 2
Section_Scale = 2.5
Detail_Scale = 0.2
Prev_Dia = 0
Next_Dia = 0
ViewType = "Fullsection"   # Fullsection, Semisection, Nosection



SP = APoint(100,100)
# End of Initial data


wb = load_workbook(r'D:\Sandeep\1_WORK IN PROGRESS\IN PROGRESS\Sandeep Jobs\AutoDraw PULSE Project\Autodraw_Shaft_Python.xlsm')  # Work Book
ws = wb.get_sheet_by_name('Datasheet')  # Work Sheet

data = []

for column in ws.iter_cols(min_row=Start_Row, max_row = End_Row, min_col=Start_Column, max_col = End_Column, values_only=True):
    data.append(column)

            
# Convert the data into a pandas DataFrame

df = pd.DataFrame(data)


# df = df.replace("", np.nan, regex=True)


df = pd.DataFrame(data, columns=['I_E', 'Dia1', 'Len', 'Tol', 'Sur_Fin', 'Is_Brg_Area', 'Is_Brg_Rel', 'Dia2', 'Is_Pin', 'Is_Spl', 'Th_Typ', 'Th_Hand', 'Th_Pit', 'Left_C_F','Left_Size','Right_C_F', 'Right_Size', 'Key_L_Sh', 'Key_R_Sh', 'Key_Width', 'Key_Depth', 'Key_Start', 'Key_Length', 'Gr_Type', 'Gr_Dia', 'Gr_Strt', 'Gr_Wid'])
df = utils.Reframe(df)
df_ext = pd.DataFrame()
df_int = pd.DataFrame()  
df_ext, df_int = utils.Split_df(df)
df_ext = utils.Extend_dataframe(df_ext)
df_int = utils.Extend_dataframe(df_int)

Coordinates = []
PointList = []
PNT1 = []
PNT2 = []
PNT3 = []
PNT4 = []
PNT5 = []
PNT6 = []
PNT7 = []
PNT8 = []

# print (df_ext)

Section_Type = "Semisection"
################################################################################################################

################################################################################################################
# Autocad instance
#
acad = Autocad(create_if_not_exists=True)

Step_SP_Ext = APoint(SP.x,SP.y)
Step_SP_Int = APoint(SP.x,SP.y)
# Step_SP = APoint(Step_SP_Ext.x,Step_SP_Ext.y)


keyPosArray_Ext, keyPosArray_Int = utils.Key_Width_List(df['I_E'],df['Len'], df['Key_Start'], df['Key_Length'], df['Key_Width'])
df_ext = utils.df_FillCoordinates(Step_SP_Ext,df_ext,"External")
df_int = utils.df_FillCoordinates(Step_SP_Int,df_int,"Internal")
# print (df_ext['GrooveCoordinates'],df_ext['GrooveRadiusStartAngle'], df_ext['GrooveRadiusEndAngle'])

# print (df_ext['GrooveCoordinates'], df_ext['GrooveRadiusStartAngle'])
keyWidth_Ext, keyWidth_Int = utils.Identify_Key_Width(keyPosArray_Ext, keyPosArray_Int, 400)


utils.DrawStep(df_int,Section_Type)
utils.DrawStep(df_ext,Section_Type)
    
        # PL = utils.DrawLine(df_ext['RightChamferCoordinates'][i], "L7")
    
    
    
    
    # # LEFT SIDE TREATMENT
    # if df['Left_C_F'][i] == "C":
    #      C1 = utils.Chamfer(Step_SP, df['Dia1'][i], df['Dia1'][i+1], df['Left_Size'][i], "Left" ,Taper_Angle, keyWidth_Ext, Section_Type,df['I_E'][i])
    #      C1.Draw_Chamfer()
        
    # elif df['Left_C_F'][i] == "R":
    #      R1 = utils.Radius(Step_SP, df['Dia1'][i], df['Dia1'][i+1], df['Left_Size'][i], "Left" ,Taper_Angle, keyWidth_Ext, Section_Type,df['I_E'][i])
    #      R1.Draw_Radius()
    # else:
    #     L1 = utils.Draw_Vertical_Edge(Step_SP, df['Dia1'][i], df['Dia1'][i+1], keyWidth_Ext, Section_Type,df['I_E'][i])
        
         
    # GROOVE
    
    # RIGHT SIDE TREATMENT


    # INCREMENT
    

#     StepEnd1 = utils.StepEnd(Step_SP,Direction,Dia1,Chamfer_Size,45,0,"0","Top")
#     StepEnd1.Draw_StepEnd()
#     StepEnd1 = utils.StepEnd(Step_SP,Direction,Dia1,Chamfer_Size,45,0,"0","Bottom")
#     StepEnd1.Draw_StepEnd()

#     Radius1 = utils.Radius(Step_SP,"Right",Radius_Size,Dia1,"0","Top")
#     Radius1.Draw_Radius() 
#     Radius1 = utils.Radius(Step_SP,"Right",Radius_Size,Dia1,"0","Top")
#     Radius1.Draw_Radius() 

    # Step_SP.x = Step_SP.x + df['Len'][i]



# L1, L2 = utils.Draw_Horizontal_Edge(APoint(10,10),60,40,50,"Fullsection","External")
# L1, L2 = utils.Draw_Vertical_Edge(APoint(10,10),40,50,4,"Fullsection","External")

# P1 = utils.Groove(APoint(10,10),"Circular",3,2,40,-5,8,"Semisection","Internal")
# L1, L2 = P1.Draw_Groove()


# P1 = utils.Radius(APoint(0,0), 40, 60, 5, "Left" ,5, "Semisection","External")
# P1.Draw_Radius()

# P1 = utils.Chamfer(APoint(0,0), 40, 60, 5, "Right" ,5, 6, "Semisection","External")
# P1.Draw_Chamfer()

# utils.DrawLine(APoint(10,10),APoint(40,40),"")
# utils.DrawArc(APoint(10,10),40,0, 90,"")


################################################################################################################
# Keyway1 = utils.Keyway(APoint(10,10),"Round","Square",8.5,67.8)
# Keyway1.Draw_Keyway()

# StepEnd1 = utils.StepEnd(APoint(10,10),"Right",40,5,45,10,"0","Top")
# StepEnd1.Draw_StepEnd()

# Groove1 = utils.Groove(APoint(10,10),"Circular",4,40,36,10)
# Groove1.Draw_Groove()

# Radius1 = utils.Radius(APoint(10,10),"Right",2,40)
# Radius1.Draw_Radius()

# Keysection1 = utils.Keysection(APoint(10,10),"Right","Open",40,8,5)
# Keysection1.Draw_Keysection()

# Text1 = utils.Insert_Text(APoint(10,10),"SUPDT.(Mech.)","Big")

# Block1 = utils.Insert_Block(APoint(10,10),"testblock",2.0,2.0,0.0,45)


# Dimlfac = 2
# Dimension1 = utils.Insert_LinDim(APoint(10,10),APoint(150,150), -20,"%%c","x45", 0, 0.4,'',Dimlfac)

# Dimension2 = utils.Insert_RadialDim(APoint(10,10),40,315,30,'',1)

# Detail_Groove1 = utils.Insert_GrooveDetail(APoint(10,10),"Circular","Internal",15)
# utils.Draw_Vertical_Edge(APoint(10,10),50,10,"Bottom","Solid")
# utils.Draw_Horizontal_Edge(APoint(0,0),"Circular",60,4,180,0,7,"0","Bottom")



################################################################################################################
